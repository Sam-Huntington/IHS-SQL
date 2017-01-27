# -*- coding: utf-8 -*-
"""
Created on Fri Jan 27 15:53:19 2017

@author: mfp48001
"""

# -*- coding: utf-8 -*-
"""
Created on Fri Jan 27 10:51:52 2017

@author: MFP48001
"""

from openpyxl import Workbook
from openpyxl import load_workbook
import pyodbc

def SQL_Connection(dbname):
    #create a conncection to the SQL server using the SVC account.  Database name needs to be changed
    cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=WAS-2PAurDB01;PORT=1433;DATABASE=RIV_H22016_v4;UID=svc_AURORAXMP;PWD=XMP')
    
    return cnxn
   
    
def create_tables(cnxn):
    #drops and recreates the NAP_data table
    cursor = cnxn.cursor()
    try:
        cursor.execute(''' drop table NAP_data''')
    except:
        pass
    
    cursor.execute('''create table NAP_data (
    
                                    NAP_Region          VARCHAR(max),
                                    Data_Category       VARCHAR(max),
                                    Data_Type           VARCHAR(max),
                                    Year                INT,
                                    Data                VARCHAR(max))''')

    cursor.commit()
    
def log_data(ret_list, cnxn):
    #uses executemany() to push the ret_list data into SQL
    cursor = cnxn.cursor()
    try:
        stmt = "INSERT INTO NAP_data (NAP_Region, Data_Category, Data_Type, Year, Data) VALUES (?,?,?,?,?)"
        cursor.executemany(stmt, ret_list)
        cursor.commit()
    except:
        pass

def parse_workbook():
    #umbrella execution function
    dbname = input('Please enter the name of the SQL database')
    filepath = input('Please enter the filepath of the retainer document')
    file = input('Please enter the filename of the retainer document')
    dbname = 'RIV_H22016_v4'
    filepath = r'C:\Users\mfp48001\Documents'
    #still working on how to fix the '\' issue, python a real bastard about parsing this
    file = r'\North American Power Market Fundamentals  Rivalry, December 2016.xlsx'
    wb = load_workbook(filename = filepath+file)
    #excel ranges defining the discrete tables in each of the NAP regional summary tabs
    tables = ['B17:AI35', 'B37:AI55','B57:AI75','B77:AI95', 'B97:AI100','B102:AI105','B108:AI126','B128:AI137','B139:AI140','B142:AI147','B150:AI166','B168:AI181','B183:AI193','B195:AI198', 'B200:AI202']
    ret_list = []
    for sheet in wb:
        if sheet.title[:2] == 'RS':
            #pull the region name
            Region_Name = sheet.cell(row=6,column=2).value[:-24]
            for table_str in tables:
                #pull the table name (top left cell in each range)
                Table_Name = sheet[table_str.split(':')[0]].value
                #call the parse function for the table
                ret_list = parse_table(sheet, table_str, ret_list, Region_Name, Table_Name)   
    cnxn = SQL_Connection(dbname)
    create_tables(cnxn)
    log_data(ret_list, cnxn)
    cnxn.close()          

def parse_table(sheet, rangestr, ret_list, Region_Name, Table_Name):
    #iterable generator
    rows = sheet.iter_rows(range_string = rangestr)
    #kick the first row as its blank except for the table name
    rows.__next__()
    for row in rows:
        year = 2008
        for cell in row[1:]:
            #parse out the data points in the table into individual rows tagged with the appropriate metadata
            out_row_list = []
            out_row_list.append(Region_Name)
            out_row_list.append(Table_Name)
            try:
                #round when value is a float
                val = round(cell.value,4)
            except:
                val = cell.value
                #row[0] is the data_type
            out_row_list.append(row[0].value)
            out_row_list.append(year)
            out_row_list.append(val)
            ret_list.append(out_row_list)
            year+=1
    return ret_list

#execute:    
parse_workbook()